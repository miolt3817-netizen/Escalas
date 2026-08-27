"""Exportação da escala — Parte 2, "Exportação".

Três formatos, cada um com um propósito diferente:

* **PDF** — o que vai para o mural do quartel. Layout de impressão, A4.
* **XLSX** — para quem precisa manipular: largura de coluna, cabeçalho fixo,
  cores por tipo de escala e uma aba de resumo por bombeiro.
* **CSV** — troca de dados com outras ferramentas. Gravado em UTF-8 **com BOM**,
  porque o Excel em português assume Windows-1252 quando não há BOM e transforma
  "terça" em "terÃ§a".

Tudo é gerado no servidor. Fazer isso no navegador foi o que produziu o CSV
ilegível: o front não controla largura de coluna nem codificação de arquivo.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MESES = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]
DIAS = [
    "segunda-feira", "terça-feira", "quarta-feira", "quinta-feira",
    "sexta-feira", "sábado", "domingo",
]

# Mesmas cores da interface (Parte 1, "Design").
MARCA = "D8181B"
CINZA_TEXTO = "5F6368"
FUNDO_BRANCA = "F1F3F4"
FUNDO_VERMELHA = "FCE8E6"
TEXTO_VERMELHA = "A50E0E"
BORDA = "DADCE0"


def _titulo(ano: int, mes: int) -> str:
    return f"{MESES[mes - 1].capitalize()} de {ano}"


def _linhas(plantoes: list[dict]) -> list[dict]:
    """Normaliza os plantões para as três saídas."""
    saida = []
    for p in sorted(plantoes, key=lambda x: x["data"]):
        d = date.fromisoformat(p["data"])
        saida.append(
            {
                "data": d,
                "data_br": d.strftime("%d/%m/%Y"),
                "dia_semana": DIAS[d.weekday()],
                "tipo": p["tipo"],
                "bombeiro": p["bombeiro"],
                "observacoes": " · ".join(
                    filter(
                        None,
                        [
                            p.get("feriado"),
                            "ajuste do supervisor" if p.get("origem") == "manual" else "",
                            "troca aprovada" if p.get("origem") == "troca" else "",
                        ],
                    )
                ),
            }
        )
    return saida


# --------------------------------------------------------------------------- #
# XLSX
# --------------------------------------------------------------------------- #


def gerar_xlsx(
    ano: int,
    mes: int,
    plantoes: list[dict],
    status: str = "rascunho",
    versao: int = 1,
    resumo: str = "",
) -> bytes:
    linhas = _linhas(plantoes)
    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"

    fina = Side(style="thin", color=BORDA)
    borda = Border(left=fina, right=fina, top=fina, bottom=fina)

    # --- título -----------------------------------------------------------
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Escala de plantão — {_titulo(ano, mes)}"
    ws["A1"].font = Font(size=14, bold=True, color="202124")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:E2")
    ws["A2"] = f"EscalaFogo · Corpo de Bombeiros · versão {versao} · {status}"
    ws["A2"].font = Font(size=10, color=CINZA_TEXTO)

    # --- cabeçalho --------------------------------------------------------
    cabecalhos = ["Data", "Dia da semana", "Tipo", "Bombeiro", "Observações"]
    for coluna, texto in enumerate(cabecalhos, start=1):
        c = ws.cell(row=4, column=coluna, value=texto)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=MARCA)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = borda
    ws.row_dimensions[4].height = 22

    # --- dados ------------------------------------------------------------
    for i, linha in enumerate(linhas, start=5):
        vermelha = linha["tipo"] == "vermelha"
        valores = [
            linha["data_br"],
            linha["dia_semana"],
            "Vermelha" if vermelha else "Branca",
            linha["bombeiro"],
            linha["observacoes"],
        ]
        for coluna, valor in enumerate(valores, start=1):
            c = ws.cell(row=i, column=coluna, value=valor)
            c.border = borda
            c.alignment = Alignment(vertical="center")
            if vermelha:
                c.fill = PatternFill("solid", fgColor=FUNDO_VERMELHA)
                if coluna == 3:
                    c.font = Font(bold=True, color=TEXTO_VERMELHA)
            elif coluna == 3:
                c.fill = PatternFill("solid", fgColor=FUNDO_BRANCA)
                c.font = Font(color=CINZA_TEXTO)

    # --- larguras: é o que faltava no CSV e produzia "########" -----------
    for coluna, largura in enumerate([13, 16, 12, 26, 30], start=1):
        ws.column_dimensions[get_column_letter(coluna)].width = largura

    ws.freeze_panes = "A5"  # cabeçalho fixo ao rolar
    ws.auto_filter.ref = f"A4:E{4 + len(linhas)}"

    # --- aba de resumo por bombeiro ---------------------------------------
    resumo_ws = wb.create_sheet("Resumo")
    contagem: dict[str, dict[str, int]] = {}
    for linha in linhas:
        c = contagem.setdefault(linha["bombeiro"], {"total": 0, "branca": 0, "vermelha": 0})
        c["total"] += 1
        c[linha["tipo"]] += 1

    resumo_ws["A1"] = f"Resumo por bombeiro — {_titulo(ano, mes)}"
    resumo_ws["A1"].font = Font(size=13, bold=True)
    resumo_ws.merge_cells("A1:D1")

    for coluna, texto in enumerate(["Bombeiro", "Total", "Branca", "Vermelha"], start=1):
        c = resumo_ws.cell(row=3, column=coluna, value=texto)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=MARCA)
        c.border = borda

    for i, (nome, c) in enumerate(sorted(contagem.items()), start=4):
        for coluna, valor in enumerate(
            [nome, c["total"], c["branca"], c["vermelha"]], start=1
        ):
            cel = resumo_ws.cell(row=i, column=coluna, value=valor)
            cel.border = borda

    for coluna, largura in enumerate([26, 10, 10, 12], start=1):
        resumo_ws.column_dimensions[get_column_letter(coluna)].width = largura

    if resumo:
        linha_resumo = 5 + len(contagem)
        resumo_ws.merge_cells(
            start_row=linha_resumo, start_column=1, end_row=linha_resumo + 3, end_column=4
        )
        c = resumo_ws.cell(row=linha_resumo, column=1, value=resumo)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=10, color=CINZA_TEXTO)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# PDF
# --------------------------------------------------------------------------- #


def _html_pdf(
    ano: int, mes: int, linhas: list[dict], status: str, versao: int, resumo: str
) -> str:
    corpo = "".join(
        f"""<tr class="{l['tipo']}">
              <td class="data">{l['data_br']}</td>
              <td>{l['dia_semana']}</td>
              <td><span class="tipo">{'Vermelha' if l['tipo'] == 'vermelha' else 'Branca'}</span></td>
              <td class="nome">{l['bombeiro']}</td>
              <td class="obs">{l['observacoes']}</td>
            </tr>"""
        for l in linhas
    )
    contagem: dict[str, int] = {}
    for l in linhas:
        contagem[l["bombeiro"]] = contagem.get(l["bombeiro"], 0) + 1
    resumo_linhas = " · ".join(
        f"{nome}: {qtd}" for nome, qtd in sorted(contagem.items())
    )

    return f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="utf-8">
<style>
/* Compactado para que um mês inteiro (até 31 dias) caiba em UMA folha:
   escala que vira duas páginas é ruim de afixar no mural. */
@page {{ size: A4; margin: 12mm 13mm 14mm;
  @bottom-center {{ content: "Página " counter(page) " de " counter(pages);
    font-family: sans-serif; font-size: 8pt; color: #80868b; }} }}
body {{ font-family: 'DejaVu Sans', sans-serif; font-size: 8.5pt; color: #202124; margin: 0; }}
h1 {{ font-size: 14pt; margin: 0 0 1mm; font-weight: 600; }}
.sub {{ font-size: 8.5pt; color: #5f6368; margin: 0 0 4mm;
  padding-bottom: 2mm; border-bottom: 1.5pt solid #d8181b; }}
table {{ width: 100%; border-collapse: collapse; }}
thead {{ display: table-header-group; }}
th {{ background: #d8181b; color: #fff; font-size: 7.5pt; text-align: left;
  padding: 1.6mm 2.2mm; font-weight: 600; text-transform: uppercase;
  letter-spacing: .04em; }}
td {{ padding: 1.4mm 2.2mm; border-bottom: .3pt solid #dadce0; }}
tr {{ page-break-inside: avoid; }}
tr.vermelha td {{ background: #fce8e6; }}
tr.vermelha .tipo {{ color: #a50e0e; font-weight: 600; }}
.tipo {{ font-size: 7.5pt; color: #5f6368; }}
.data {{ font-variant-numeric: tabular-nums; white-space: nowrap; }}
.nome {{ font-weight: 500; }}
.obs {{ font-size: 7.5pt; color: #5f6368; }}
.rodape {{ margin-top: 4mm; padding-top: 2.5mm; border-top: .3pt solid #dadce0;
  font-size: 7.5pt; color: #5f6368; line-height: 1.45; }}
</style></head><body>
<h1>Escala de plantão — {_titulo(ano, mes)}</h1>
<p class="sub">EscalaFogo &nbsp;·&nbsp; Corpo de Bombeiros &nbsp;·&nbsp; versão {versao} &nbsp;·&nbsp; {status}</p>
<table>
  <thead><tr><th>Data</th><th>Dia</th><th>Tipo</th><th>Bombeiro</th><th>Observações</th></tr></thead>
  <tbody>{corpo}</tbody>
</table>
<div class="rodape">
  <b>Plantões no mês:</b> {resumo_linhas}
  {f'<br><br>{resumo}' if resumo else ''}
</div>
</body></html>"""


def _pdf_reportlab(
    ano: int, mes: int, linhas: list[dict], status: str, versao: int, resumo: str
) -> bytes:
    """Alternativa em ReportLab, usada quando o WeasyPrint não está disponível.

    O WeasyPrint depende de bibliotecas GTK nativas, que não existem no Windows
    sem instalação separada — por isso o aplicativo desktop usa este caminho.
    O ReportLab é Python puro e empacota sem atrito.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    MARCA_RL = colors.HexColor("#d8181b")
    CINZA_RL = colors.HexColor("#5f6368")
    VERMELHO_FUNDO = colors.HexColor("#fce8e6")
    VERMELHO_TEXTO = colors.HexColor("#a50e0e")
    LINHA = colors.HexColor("#dadce0")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=13 * mm, rightMargin=13 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=f"Escala de plantão — {_titulo(ano, mes)}",
    )

    estilo_titulo = ParagraphStyle(
        "titulo", fontName="Helvetica-Bold", fontSize=14,
        textColor=colors.HexColor("#202124"), spaceAfter=2, alignment=TA_LEFT,
    )
    estilo_sub = ParagraphStyle(
        "sub", fontName="Helvetica", fontSize=8.5, textColor=CINZA_RL, spaceAfter=6,
    )
    estilo_rodape = ParagraphStyle(
        "rodape", fontName="Helvetica", fontSize=7.5, textColor=CINZA_RL, leading=11,
    )
    estilo_celula = ParagraphStyle(
        "celula", fontName="Helvetica", fontSize=8, leading=10,
    )

    elementos = [
        Paragraph(f"Escala de plantão — {_titulo(ano, mes)}", estilo_titulo),
        Paragraph(
            f"EscalaFogo &nbsp;·&nbsp; Corpo de Bombeiros &nbsp;·&nbsp; "
            f"versão {versao} &nbsp;·&nbsp; {status}",
            estilo_sub,
        ),
        Spacer(1, 3 * mm),
    ]

    dados = [["Data", "Dia", "Tipo", "Bombeiro", "Observações"]]
    for l in linhas:
        dados.append(
            [
                l["data_br"],
                l["dia_semana"],
                "Vermelha" if l["tipo"] == "vermelha" else "Branca",
                l["bombeiro"],
                Paragraph(l["observacoes"], estilo_celula) if l["observacoes"] else "",
            ]
        )

    tabela = Table(
        dados,
        colWidths=[24 * mm, 27 * mm, 20 * mm, 46 * mm, 67 * mm],
        repeatRows=1,
    )
    estilos = [
        ("BACKGROUND", (0, 0), (-1, 0), MARCA_RL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, LINHA),
    ]
    for i, l in enumerate(linhas, start=1):
        if l["tipo"] == "vermelha":
            estilos.append(("BACKGROUND", (0, i), (-1, i), VERMELHO_FUNDO))
            estilos.append(("TEXTCOLOR", (2, i), (2, i), VERMELHO_TEXTO))
            estilos.append(("FONTNAME", (2, i), (2, i), "Helvetica-Bold"))
        else:
            estilos.append(("TEXTCOLOR", (2, i), (2, i), CINZA_RL))
    tabela.setStyle(TableStyle(estilos))
    elementos.append(tabela)

    contagem: dict[str, int] = {}
    for l in linhas:
        contagem[l["bombeiro"]] = contagem.get(l["bombeiro"], 0) + 1
    texto_resumo = " · ".join(f"{n}: {q}" for n, q in sorted(contagem.items()))

    elementos.append(Spacer(1, 4 * mm))
    elementos.append(Paragraph(f"<b>Plantões no mês:</b> {texto_resumo}", estilo_rodape))
    if resumo:
        elementos.append(Spacer(1, 2 * mm))
        elementos.append(Paragraph(resumo, estilo_rodape))

    def rodape(canvas, documento):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(CINZA_RL)
        canvas.drawCentredString(
            A4[0] / 2, 8 * mm, f"Página {documento.page}"
        )
        canvas.restoreState()

    doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def gerar_pdf(
    ano: int,
    mes: int,
    plantoes: list[dict],
    status: str = "rascunho",
    versao: int = 1,
    resumo: str = "",
) -> bytes:
    """Usa WeasyPrint quando disponível; senão, ReportLab.

    Os dois produzem o mesmo layout. A escolha é só de disponibilidade: o
    servidor (Docker/Linux) tem as bibliotecas do WeasyPrint; o aplicativo
    desktop no Windows, não.
    """
    linhas = _linhas(plantoes)
    try:
        from weasyprint import HTML  # noqa: PLC0415 - opcional, carrega nativas

        html = _html_pdf(ano, mes, linhas, status, versao, resumo)
        return HTML(string=html).write_pdf()
    except Exception:  # noqa: BLE001 - ausência ou falha de biblioteca nativa
        return _pdf_reportlab(ano, mes, linhas, status, versao, resumo)


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #


def gerar_csv(ano: int, mes: int, plantoes: list[dict]) -> bytes:
    """UTF-8 **com BOM** e separador `;`.

    Sem o BOM, o Excel em português lê o arquivo como Windows-1252 e exibe
    "terÃ§a" no lugar de "terça". O separador `;` é o esperado em configuração
    regional pt-BR, onde a vírgula é separador decimal.
    """
    import csv

    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    escritor.writerow(["Data", "Dia da semana", "Tipo", "Bombeiro", "Observações"])
    for l in _linhas(plantoes):
        escritor.writerow(
            [
                l["data_br"],
                l["dia_semana"],
                "Vermelha" if l["tipo"] == "vermelha" else "Branca",
                l["bombeiro"],
                l["observacoes"],
            ]
        )
    return buffer.getvalue().encode("utf-8-sig")


NOMES_ARQUIVO = {
    "pdf": ("escala-{ano}-{mes:02d}.pdf", "application/pdf"),
    "xlsx": (
        "escala-{ano}-{mes:02d}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
    "csv": ("escala-{ano}-{mes:02d}.csv", "text/csv; charset=utf-8"),
}
