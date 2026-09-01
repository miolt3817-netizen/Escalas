"""Testes de layout e exportação.

O layout roda em navegador de verdade (Playwright) porque os defeitos que
apareceram — cabeçalho de tabela flutuando sobre as linhas, botões cortados,
página rolando na horizontal — são invisíveis para teste de API. Só aparecem
quando o CSS é aplicado.

Se o Playwright ou o Chromium não estiverem disponíveis, os testes de layout
são pulados; os de exportação rodam sempre.
"""

from __future__ import annotations

from datetime import date, timedelta

import pytest

from api import exportacao

def _token_layout(cliente):
    r = cliente.post(
        "/auth/login",
        data={"username": "admin@cb.gov.br", "password": "senha123"},
    )
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


NOMES = [
    "João Batista", "Maria Aparecida", "Carlos Eduardo", "Ana Paula",
    "Pedro Henrique", "Lúcia Helena", "Rafael Souza", "Beatriz Lima",
]


def _plantoes(ano: int = 2026, mes: int = 10, dias: int = 31) -> list[dict]:
    saida = []
    for i in range(dias):
        d = date(ano, mes, 1) + timedelta(days=i)
        vermelha = d.weekday() >= 5
        saida.append(
            {
                "data": d.isoformat(),
                "tipo": "vermelha" if vermelha else "branca",
                "bombeiro": NOMES[i % len(NOMES)],
                "bombeiro_id": i % len(NOMES) + 1,
                "feriado": "Feriado Teste" if i == 11 else None,
                "origem": "manual" if i == 20 else "solver",
                "travado": i == 20,
            }
        )
    return saida


# --------------------------------------------------------------------------- #
# Exportação
# --------------------------------------------------------------------------- #


def test_csv_tem_bom_e_acentos_corretos():
    """Sem BOM, o Excel em português lê como Windows-1252 e mostra 'terÃ§a'."""
    conteudo = exportacao.gerar_csv(2026, 10, _plantoes())

    assert conteudo.startswith(b"\xef\xbb\xbf"), "faltou o BOM de UTF-8"
    texto = conteudo.decode("utf-8-sig")
    assert "terça-feira" in texto
    assert "sábado" in texto
    assert "João Batista" in texto
    assert "Ã" not in texto, "acentuação quebrada"
    assert texto.splitlines()[0].count(";") == 4  # separador pt-BR


def test_csv_usa_data_brasileira():
    texto = exportacao.gerar_csv(2026, 10, _plantoes()).decode("utf-8-sig")
    assert "01/10/2026" in texto
    assert "2026-10-01" not in texto


def test_xlsx_tem_larguras_definidas():
    """Coluna estreita demais é o que produzia '########' no Excel."""
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(exportacao.gerar_xlsx(2026, 10, _plantoes())))
    assert wb.sheetnames == ["Escala", "Resumo"]

    ws = wb["Escala"]
    for coluna in "ABCDE":
        largura = ws.column_dimensions[coluna].width
        assert largura and largura >= 10, f"coluna {coluna} sem largura definida"

    assert ws.freeze_panes == "A5"  # cabeçalho fixo ao rolar
    assert ws.auto_filter.ref
    assert [ws.cell(4, c).value for c in range(1, 6)] == [
        "Data", "Dia da semana", "Tipo", "Bombeiro", "Observações"
    ]
    assert ws.cell(5, 1).value == "01/10/2026"


def test_xlsx_marca_escala_vermelha():
    import io

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(exportacao.gerar_xlsx(2026, 10, _plantoes())))["Escala"]
    tipos = {}
    for linha in range(5, 36):
        tipos[ws.cell(linha, 1).value] = (
            ws.cell(linha, 3).value,
            ws.cell(linha, 1).fill.fgColor.rgb,
        )
    # 03/10/2026 é sábado
    tipo, fundo = tipos["03/10/2026"]
    assert tipo == "Vermelha"
    assert "FCE8E6" in str(fundo)


def test_xlsx_tem_aba_de_resumo():
    import io

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(exportacao.gerar_xlsx(2026, 10, _plantoes())))["Resumo"]
    assert [ws.cell(3, c).value for c in range(1, 5)] == [
        "Bombeiro", "Total", "Branca", "Vermelha"
    ]
    total = sum(ws.cell(l, 2).value or 0 for l in range(4, 20))
    assert total == 31


def test_pdf_cabe_em_uma_pagina():
    """Mês inteiro em duas folhas é ruim de afixar no mural."""
    pdf = pytest.importorskip("weasyprint") and exportacao.gerar_pdf(
        2026, 10, _plantoes(), "publicada", 1, "Resumo de teste."
    )
    assert pdf.startswith(b"%PDF")
    # cada página gera um objeto /Type /Page
    assert pdf.count(b"/Type /Page\n") <= 1 or pdf.count(b"/Count 1") >= 1


def test_formato_invalido_e_recusado():
    assert "docx" not in exportacao.NOMES_ARQUIVO
    assert set(exportacao.NOMES_ARQUIVO) == {"pdf", "xlsx", "csv"}


# --------------------------------------------------------------------------- #
# Layout — navegador de verdade
# --------------------------------------------------------------------------- #

DETECTOR = """
() => {
  const problemas = [];
  const doc = document.documentElement;
  if (doc.scrollWidth > doc.clientWidth + 2) {
    problemas.push(`página rola na horizontal: ${doc.scrollWidth} > ${doc.clientWidth}`);
  }
  document.querySelectorAll('*').forEach(el => {
    if (el.offsetParent === null) return;
    const r = el.getBoundingClientRect();
    if (r.width && r.right > doc.clientWidth + 3 &&
        !el.closest('.tabela-rolavel') && !el.closest('.abas') &&
        !el.closest('.lateral-menu')) {
      problemas.push(`transborda: <${el.tagName.toLowerCase()}> right=${Math.round(r.right)}`);
    }
  });
  document.querySelectorAll('.tabela-rolavel').forEach((c, i) => {
    if (c.scrollWidth > c.clientWidth + 2) {
      problemas.push(`tabela ${i} cortada: ${c.scrollWidth} em ${c.clientWidth}`);
    }
  });
  document.querySelectorAll('table').forEach((t, i) => {
    const th = t.querySelector('thead th'), td = t.querySelector('tbody tr td');
    if (!th || !td) return;
    const a = th.getBoundingClientRect(), b = td.getBoundingClientRect();
    if (a.left < -1000 || a.width === 0) return;
    if (a.bottom > b.top + 2 && a.top < b.bottom) {
      problemas.push(`tabela ${i}: cabeçalho sobre a 1ª linha`);
    }
  });
  return [...new Set(problemas)];
}
"""


@pytest.mark.parametrize(
    "largura,altura,rotulo",
    [(1440, 900, "desktop"), (820, 1180, "tablet"), (390, 844, "celular")],
)
def test_layout_sem_defeitos(largura, altura, rotulo, servidor_web):
    """Nenhuma tela pode transbordar, cortar conteúdo ou sobrepor cabeçalho."""
    sync_playwright = pytest.importorskip(
        "playwright.sync_api", reason="Playwright não instalado"
    ).sync_playwright

    with sync_playwright() as pw:
        try:
            navegador = pw.chromium.launch()
        except Exception as erro:  # noqa: BLE001
            pytest.skip(f"Chromium indisponível: {erro}")

        page = navegador.new_page(viewport={"width": largura, "height": altura})
        page.goto(servidor_web, wait_until="networkidle")

        for senha in ("bombeiros2026", "SenhaDeTeste2026"):
            page.fill("#email", "supervisor@cb.sc.gov.br")
            page.fill("#senha", senha)
            page.click("text=Entrar")
            page.wait_for_timeout(1500)
            if page.is_visible("#tela-senha"):
                page.fill("#senha-nova", "SenhaDeTeste2026")
                page.fill("#senha-conf", "SenhaDeTeste2026")
                page.click("text=Salvar e continuar")
                page.wait_for_timeout(1800)
            if page.is_visible("#app"):
                break

        assert page.is_visible("#app"), "não entrou na aplicação"

        def abrirGaveta():
            """Em tela estreita a barra de abas rola na horizontal: a aba
            precisa entrar em vista antes de ser clicável."""
            page.evaluate("() => document.getElementById('abas')?.scrollTo(0, 0)")
            page.wait_for_timeout(150)

        def irAba(vista):
            page.evaluate(
                f"""() => document.querySelector('[data-vista="{vista}"]')
                        ?.scrollIntoView({{block:'nearest', inline:'center'}})"""
            )
            page.wait_for_timeout(250)
            page.click(f'[data-vista="{vista}"]')

        falhas: list[str] = []
        # "Equipe" e "Gerar escala" viraram sub-abas de Configurações; as
        # duas precisam ser abertas para o detector alcançar seu conteúdo.
        for vista, subs in [("escala", []), ("datas", []),
                            ("config", ["equipe", "gerar"]),
                            ("relatorios", [])]:
            irAba(vista)
            page.wait_for_timeout(900)
            for sub in subs or [None]:
                if sub:
                    page.click(f'#vista-{vista} .sub-aba[data-sub="{sub}"]')
                    page.wait_for_timeout(800)
                nome = f"{vista}/{sub}" if sub else vista
                for posicao in (0, 900):
                    page.evaluate(f"window.scrollTo(0, {posicao})")
                    page.wait_for_timeout(300)
                    falhas += [
                        f"{rotulo}/{nome}: {p}" for p in page.evaluate(DETECTOR)
                    ]

        navegador.close()

    assert not falhas, "\n".join(sorted(set(falhas)))


# --------------------------------------------------------------------------- #
# Aplicativo desktop
# --------------------------------------------------------------------------- #


def test_pdf_reportlab_equivale_ao_weasyprint():
    """O aplicativo Windows gera PDF sem WeasyPrint (que precisa de GTK).

    O ReportLab é o caminho alternativo. Precisa produzir o mesmo documento:
    uma página, com os dados todos.
    """
    pytest.importorskip("reportlab")
    linhas = exportacao._linhas(_plantoes())
    pdf = exportacao._pdf_reportlab(2026, 10, linhas, "publicada", 1, "Resumo.")

    assert pdf.startswith(b"%PDF")
    assert b"/Count 1" in pdf, "mês inteiro deveria caber em uma página"
    assert len(pdf) > 1500


def test_gerar_pdf_cai_no_reportlab_sem_weasyprint(monkeypatch):
    """Se o WeasyPrint faltar ou falhar, o PDF ainda tem que sair."""
    import builtins

    real = builtins.__import__

    def sem_weasyprint(nome, *args, **kwargs):
        if nome.startswith("weasyprint"):
            raise ImportError("simulando ausência do WeasyPrint")
        return real(nome, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", sem_weasyprint)
    pdf = exportacao.gerar_pdf(2026, 10, _plantoes(), "publicada", 1, "")
    assert pdf.startswith(b"%PDF")


def test_desktop_configura_ambiente_isolado(tmp_path, monkeypatch):
    """Pasta de dados própria, chave estável e porta livre."""
    import desktop

    monkeypatch.setenv("XDG_DATA_HOME", str(tmp_path))
    monkeypatch.setenv("APPDATA", str(tmp_path))

    pasta = desktop.pasta_de_dados()
    assert pasta.exists()
    assert pasta.name == "EscalasBombeiros"

    # A chave precisa sobreviver entre execuções: se mudasse, todo mundo
    # seria deslogado a cada abertura.
    primeira = desktop.segredo_persistente(pasta)
    assert len(primeira) >= 40
    assert desktop.segredo_persistente(pasta) == primeira

    porta = desktop.porta_livre()
    assert 1 <= porta <= 65535


def test_desktop_acha_porta_alternativa_se_ocupada():
    import socket

    import desktop

    with socket.socket() as ocupada:
        ocupada.bind(("0.0.0.0", 0))
        tomada = ocupada.getsockname()[1]
        ocupada.listen(1)
        alternativa = desktop.porta_livre(tomada)
        assert alternativa != tomada, "deveria escolher outra porta"


# --------------------------------------------------------------------------- #
# Aplicativo instalável (PWA)
# --------------------------------------------------------------------------- #


def test_manifest_e_valido(servidor_web):
    """Sem estes campos o navegador não oferece 'Instalar'."""
    import json
    import urllib.request

    with urllib.request.urlopen(servidor_web + "/manifest.json", timeout=20) as r:
        assert r.headers.get("Content-Type", "").startswith(
            "application/manifest+json"
        )
        manifesto = json.loads(r.read())

    assert {"name", "short_name", "start_url", "display", "icons"} <= set(manifesto)
    assert manifesto["display"] == "standalone", "abriria com a barra do navegador"
    assert manifesto["theme_color"].startswith("#")

    tamanhos = {i["sizes"] for i in manifesto["icons"]}
    assert {"192x192", "512x512"} <= tamanhos
    # O Android recorta o ícone em círculo; sem 'maskable' o desenho fica torto.
    assert any(i.get("purpose") == "maskable" for i in manifesto["icons"])

    # Todo ícone declarado precisa existir de verdade: manifest apontando para
    # arquivo ausente faz o navegador recusar a instalação sem dizer por quê.
    for icone in manifesto["icons"]:
        with urllib.request.urlopen(servidor_web + icone["src"], timeout=20) as r:
            assert r.status == 200, f"ícone ausente: {icone['src']}"
            assert len(r.read()) > 500


def test_arquivos_do_pwa_sao_servidos(servidor_web):
    import urllib.request

    for rota, tipo in [
        ("/sw.js", "application/javascript"),
        ("/estatico/marca-192-v2.png", "image/png"),
        ("/estatico/marca-512-v2.png", "image/png"),
        ("/estatico/marca-apple-v2.png", "image/png"),
    ]:
        with urllib.request.urlopen(servidor_web + rota, timeout=20) as r:
            assert r.status == 200
            assert r.headers.get("Content-Type", "").startswith(tipo)
            assert len(r.read()) > 100


def test_rota_de_estaticos_resiste_a_travessia(servidor_web):
    """`/estatico/` não pode virar porta de saída para o resto do disco."""
    import urllib.error
    import urllib.request

    for tentativa in (
        "/estatico/../api/main.py",
        "/estatico/..%2f..%2fapi%2fmain.py",
        "/estatico/inexistente.png",
    ):
        try:
            codigo = urllib.request.urlopen(
                servidor_web + tentativa, timeout=10
            ).status
        except urllib.error.HTTPError as erro:
            codigo = erro.code
        assert codigo == 404, f"{tentativa} devolveu {codigo}"


def test_pagina_nao_depende_de_servidor_externo(servidor_web):
    """Aplicativo instalado precisa funcionar sem internet.

    Também é questão de privacidade: buscar fonte em CDN externa envia o IP de
    cada usuário a um terceiro, o que não se justifica num sistema que lida com
    dado de saúde.
    """
    import urllib.request

    html = urllib.request.urlopen(servidor_web + "/", timeout=20).read().decode()
    for externo in ("fonts.googleapis.com", "fonts.gstatic.com", "cdn.", "unpkg.com"):
        assert externo not in html, f"depende de {externo}"


def test_service_worker_nao_guarda_dados(servidor_web):
    """A casca pode ficar em cache; os dados, não.

    Escala guardada no aparelho fica desatualizada em silêncio — alguém pode
    aparecer para trabalhar no dia errado. E indisponibilidade envolve dado de
    saúde, que não deve sobrar no disco.
    """
    sync_playwright = pytest.importorskip(
        "playwright.sync_api", reason="Playwright não instalado"
    ).sync_playwright

    with sync_playwright() as pw:
        try:
            navegador = pw.chromium.launch()
        except Exception as erro:  # noqa: BLE001
            pytest.skip(f"Chromium indisponível: {erro}")

        page = navegador.new_page(viewport={"width": 390, "height": 844})
        page.goto(servidor_web, wait_until="networkidle")
        page.wait_for_timeout(2500)

        registro = page.evaluate(
            "async () => { const r = await navigator.serviceWorker.getRegistration();"
            " return r ? r.active?.state : null; }"
        )
        assert registro == "activated", "service worker não ativou"

        for senha in ("bombeiros2026", "SenhaDeTeste2026"):
            page.fill("#email", "supervisor@cb.sc.gov.br")
            page.fill("#senha", senha)
            page.click("text=Entrar")
            page.wait_for_timeout(1400)
            if page.is_visible("#tela-senha"):
                page.fill("#senha-nova", "SenhaDeTeste2026")
                page.fill("#senha-conf", "SenhaDeTeste2026")
                page.click("text=Salvar e continuar")
                page.wait_for_timeout(1600)
            if page.is_visible("#app"):
                break
        page.wait_for_timeout(2500)

        guardado = page.evaluate(
            "async () => { const n = await caches.keys();"
            " if (!n.length) return [];"
            " const c = await caches.open(n[0]); const k = await c.keys();"
            " return k.map(r => new URL(r.url).pathname); }"
        )
        navegador.close()

    dados = [
        u for u in guardado
        if any(u.startswith(p) for p in (
            "/auth", "/escalas", "/usuarios", "/equidade",
            "/indisponibilidades", "/preferencias", "/jobs", "/auditoria",
        ))
    ]
    assert not dados, f"dados de API em cache: {dados}"
    assert "/" in guardado, "a casca deveria estar em cache para abrir offline"


def test_sessao_sobrevive_ao_f5(servidor_web):
    """Recarregar a página não pode derrubar a sessão.

    A sessão fica no localStorage e é revalidada contra o servidor a cada
    carregamento: token vencido, forjado ou de usuário desativado cai no login.
    """
    sync_playwright = pytest.importorskip(
        "playwright.sync_api", reason="Playwright não instalado"
    ).sync_playwright

    with sync_playwright() as pw:
        try:
            navegador = pw.chromium.launch()
        except Exception as erro:  # noqa: BLE001
            pytest.skip(f"Chromium indisponível: {erro}")

        contexto = navegador.new_context(viewport={"width": 1280, "height": 800})
        page = contexto.new_page()
        page.goto(servidor_web, wait_until="networkidle")

        for senha in ("bombeiros2026", "SenhaDeTeste2026"):
            page.fill("#email", "supervisor@cb.sc.gov.br")
            page.fill("#senha", senha)
            page.click("text=Entrar")
            page.wait_for_timeout(1500)
            if page.is_visible("#tela-senha"):
                page.fill("#senha-nova", "SenhaDeTeste2026")
                page.fill("#senha-conf", "SenhaDeTeste2026")
                page.click("text=Salvar e continuar")
                page.wait_for_timeout(1800)
            if page.is_visible("#app"):
                break
        assert page.is_visible("#app"), "não entrou"

        # F5
        page.reload(wait_until="networkidle")
        page.wait_for_timeout(2500)
        assert page.is_visible("#app"), "F5 derrubou a sessão"
        assert not page.is_visible("#tela-login")

        # aba nova no mesmo navegador
        outra = contexto.new_page()
        outra.goto(servidor_web, wait_until="networkidle")
        outra.wait_for_timeout(2500)
        assert outra.is_visible("#app"), "aba nova deveria aproveitar a sessão"
        outra.close()

        # sair limpa tudo
        page.click("#btn-sair")
        page.wait_for_timeout(2500)
        assert page.is_visible("#tela-login")
        assert page.evaluate("() => localStorage.getItem('escalas.sessao')") is None
        page.reload(wait_until="networkidle")
        page.wait_for_timeout(2000)
        assert page.is_visible("#tela-login"), "voltou logado depois de sair"

        # token forjado é recusado pelo servidor
        page.evaluate(
            "() => localStorage.setItem('escalas.sessao',"
            " JSON.stringify({token:'forjado.invalido.xyz',"
            " papel:'administrador', nome:'Intruso'}))"
        )
        page.reload(wait_until="networkidle")
        page.wait_for_timeout(2500)
        assert page.is_visible("#tela-login"), "token forjado entrou"
        assert page.evaluate("() => localStorage.getItem('escalas.sessao')") is None

        navegador.close()


def test_javascript_da_interface_e_valido():
    """Sintaxe conferida pelo Node, que suporta o mesmo ES do navegador."""
    import re
    import shutil
    import subprocess
    import tempfile
    from pathlib import Path

    if not shutil.which("node"):
        pytest.skip("Node não instalado")

    html = (Path(__file__).resolve().parent.parent / "web" / "index.html").read_text()
    blocos = re.findall(r"<script>(.*?)</script>", html, re.S)
    assert blocos, "nenhum bloco de script encontrado"

    for i, bloco in enumerate(blocos):
        with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False) as arq:
            arq.write(bloco)
            caminho = arq.name
        r = subprocess.run(
            ["node", "--check", caminho], capture_output=True, text=True, timeout=60
        )
        assert r.returncode == 0, f"bloco {i}: {r.stderr[:400]}"

    # toda função chamada por onclick/onchange precisa existir
    chamadas = set(re.findall(r'on(?:click|change)="(\w+)\(', html)) - {"if"}
    definidas = set(re.findall(r"(?:async )?function (\w+)", html))
    assert not chamadas - definidas, f"sem definição: {chamadas - definidas}"

    # todo getElementById precisa apontar para um id que existe
    usados = set(re.findall(r"getElementById\('([\w-]+)'\)", html))
    existentes = set(re.findall(r'id="([\w-]+)"', html))
    assert not usados - existentes, f"ids inexistentes: {usados - existentes}"


def test_sem_dialogos_nativos_do_navegador():
    """confirm() e prompt() travam a página e ignoram o visual do sistema.

    No celular ainda aparecem com o endereço do site em cima, o que num sistema
    de trabalho passa impressão de improviso.
    """
    import re
    from pathlib import Path

    html = (Path(__file__).resolve().parent.parent / "web" / "index.html").read_text()
    js = "\n".join(re.findall(r"<script>(.*?)</script>", html, re.S))
    js = re.sub(r"/\*.*?\*/", "", js, flags=re.S)  # ignora comentários

    for nativo in (r"\bconfirm\s*\(", r"(?<!promptInstalacao\.)\bprompt\s*\(",
                   r"\balert\s*\("):
        assert not re.search(nativo, js), f"uso de diálogo nativo: {nativo}"


def test_modal_de_confirmacao_fica_acima_do_painel_do_dia():
    """Regressão: com o mesmo z-index, o painel do dia interceptava os cliques
    e o botão de confirmar ficava inalcançável."""
    import re
    from pathlib import Path

    html = (Path(__file__).resolve().parent.parent / "web" / "index.html").read_text()
    base = re.search(r"\.modal\{[^}]*z-index:\s*(\d+)", html)
    acima = re.search(r"#modal-confirmar\{[^}]*z-index:\s*(\d+)", html)
    assert base and acima, "z-index não declarado"
    assert int(acima.group(1)) > int(base.group(1))
